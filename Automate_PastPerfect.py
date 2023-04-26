#Connor Johnson
#Selenium bot to automate the photo portion inputs of data given by an excel sheet to the datebase: PastPerfect
#Good Chrome extension to find paths in website: SelectorsHub

#Make sure to pip install: selenium, webdriver-manager, and openpyxl
#
#Mac-->Windows
#change: Chromedriver path, photodrive path
#

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from datetime import datetime
import logging
import time
import os.path

logging.basicConfig(filename='test.log', filemode='w', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

#Mac path
# PATH = "/Users/Johnson_code/chromedriver"
#Windows path
PATH = "C:\AutomationTool\chromedriver"

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()
driver.get("https://mypastperfect.com/Account/Login?ReturnUrl=%2F")     #opens browser to website

username = driver.find_element(By.ID, 'Email')
username.send_keys("connor.johnson1@marist.edu")                        #finds unsername

password = driver.find_element(By.ID, 'Password')
password.send_keys("!Lacrosse5647")
password.send_keys(Keys.RETURN)                                         #finds password and clicks enter to login

                                                                        #use the try-catch for it to wait until it finds the page


            ###--------------Page Navigation--------------###
try:
    catalog_click = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "catalogs-home-button"))
    )
    catalog_click.click()#clicks the catalog button

    photograph_click = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="object-type-tabs"]/div/ul/li[2]/a'))
    )
    photograph_click.click()#clicks the photograph button

except:
    print("Could not find button")
    #logging.error()


            ###--------------Excel Import--------------###
book = load_workbook('MAPC Series 2 Buildings and Grounds.xlsx')  # Excel file from personal computer
sheet = book.active                                                 #reads excel file inside folder
row_num = 4

while row_num <= sheet.max_row:
   for cell in sheet[row_num]:

        excel_Identifier = sheet['B' + str(row_num)].value                                #not using value yet, only tests
        excel_Title = sheet['C' + str(row_num)].value
        excel_Author = sheet['D' + str(row_num)].value
        excel_Description = sheet['E' + str(row_num)].value
        excel_Date = sheet['F' + str(row_num)].value
        strDate = str(excel_Date)
        excel_Object = sheet['G' + str(row_num)].value
        excel_Format = sheet['H' + str(row_num)].value
        excel_Subject = sheet['I' + str(row_num)].value
        excel_Collection = sheet['J' + str(row_num)].value
        excel_Media = sheet['K' + str(row_num)].value
        excel_Relation = sheet['L' + str(row_num)].value


                    ###--------------Insert New Record--------------###

        try:
            add_record_click = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="catalogs-grid"]/div/div[2]/div[1]/div/button[1]'))
            )
            add_record_click.click()#clicks the 'add record' button

        except:
            print("Could not add new record")

        time.sleep(2)
        ObjectID_New_Record = driver.find_element(By.XPATH, '//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[4]/div/input')
        ObjectID_New_Record.send_keys(excel_Identifier)  #identifier                    #finds Object ID Box and inserts from excel sheet

        Search_ObjectName_New_Record = driver.find_element(By.XPATH, '//*[@id="newRecordNameLex"]/div/div/span[2]')
        Search_ObjectName_New_Record.click()                           #finds the search icon to go find Object Name

        time.sleep(2)
        if excel_Object.lower().strip() == 'photograph':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Select_New_Record = Select(driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[4]/div[3]/select'))
            Select_New_Record.select_by_visible_text('20')            #Selects the dropdown to make list bigger

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[3]/table/tbody/tr[13]/td[1]')
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'Photograph' object name

        if excel_Object.lower().strip() == 'slide':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Select_New_Record = Select(driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[4]/div[3]/select'))
            Select_New_Record.select_by_visible_text('20')            #Selects the dropdown to make list bigger

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, '/html[1]/body[1]/div[8]/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[10]/td[1]')
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'slide' object name

        if excel_Object.lower().strip() == 'dvd':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object.lower().strip())                        #Inserts object from excel into search

            time.sleep(2)
            Select_New_Record = Select(driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[4]/div[3]/select'))
            Select_New_Record.select_by_visible_text('20')            #Selects the dropdown to make list bigger

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, "//td[normalize-space()='DVD']")
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'dvd' object name


        if excel_Object.lower().strip() == 'print':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Select_New_Record = Select(driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[4]/div[3]/select'))
            Select_New_Record.select_by_visible_text('20')            #Selects the dropdown to make list bigger

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, '/html[1]/body[1]/div[8]/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[18]/td[1]')
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'print' object name

        if excel_Object.lower().strip() == 'negative':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[3]/table/tbody/tr[2]/td[1]')
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'negative' object name

        if excel_Object.lower().strip() == 'postcard':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[3]/table/tbody/tr[4]/td[1]') #insert xpath
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'post card' object name

        if excel_Object.lower().strip() == 'cd':
            ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
            ObjectName_New_Record.send_keys(excel_Object)                        #Inserts object from excel into search

            time.sleep(2)
            Select_New_Record = Select(driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[4]/div[2]/div[4]/div[3]/select'))
            Select_New_Record.select_by_visible_text('20')            #Selects the dropdown to make list bigger

            time.sleep(2)
            Action = ActionChains(driver)
            Action_ObjectName_New_Record = driver.find_element(By.XPATH, ) #instert xpath
            Action.double_click(Action_ObjectName_New_Record).perform()   #finds the 'cd' object name

        time.sleep(2)
        Title_New_Record = driver.find_element(By.XPATH, '//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[6]/div/input')
        Title_New_Record.send_keys(excel_Title)                              #Inserts title from excel

        if excel_Description.lower() != 'none':
            Description_New_Record = driver.find_element(By.XPATH, '//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[7]/div/textarea')
            Description_New_Record.send_keys(excel_Description)                  #Inserts description from excel


        time.sleep(2)
        Add_New_Record = driver.find_element(By.XPATH, '//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[8]/button[1]')
        Add_New_Record.click()                              #clicks the 'add new record' button
        time.sleep(2)

                        ###--------------Going to second webpage after this point--------------###


                        ###--------------Image Management--------------###

        time.sleep(2)

        other_Name_Click = driver.find_element(By.XPATH, '//*[@id="otherNameDictionary"]/div/span')
        other_Name_Click.click()

        time.sleep(2)

        other_Name_Click = driver.find_element(By.XPATH, "//td[normalize-space()='Poughkeepsie Community and Surroundings']")
        other_Name_Click.click()

        time.sleep(2)

        #Mac Path
        #if os.path.exists("/Volumes/Transfer2/Marist Brothers NEW/" + excel_Media):

        #Windows path
        if os.path.exists(r"D:\Buildings and Grounds NEW\\" + excel_Media):
            try:
                Enter_Image_Managemnet = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.menu-btn-primary.col-md-8'))
                )
                Enter_Image_Managemnet.click()

                Add_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
                )
                Add_Image_Management.send_keys(
                    r"D:\Buildings and Grounds NEW\\" + excel_Media)

                time.sleep(6)
                Checkbox_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "input[data-bind='checked: IsPublic, enable: $parent.editMode()']"))
                )
                Checkbox_Image_Management.click()

                Save_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "button[data-bind='click: save, enable: isSaveButtonEnabled()']"))
                )
                Save_Image_Management.click()
            except:
                print("Could not finish Image Management")
        else:
            print("File not found: " + excel_Identifier)


        try:

            currentDate = time.strftime("%m/%d/%Y")
            Catalog_Date_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="catalogDate"]/div/input'))
            )
            Catalog_Date_Edit.send_keys(currentDate)

            if len(strDate) == 19 :
                # Convert the string to a datetime object
                date_obj = datetime.strptime(strDate, "%Y-%m-%d %H:%M:%S")

                # Format the datetime object to the desired output format
                output_str = date_obj.strftime("%Y %B %d")

                Date_Edit = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="Date"]'))
                )
                Date_Edit.send_keys(output_str)

            else:
                Date_Edit = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="Date"]'))
                )
                Date_Edit.send_keys(excel_Date)

            Public_Access_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'public-access-checkbox'))
            )
            Public_Access_Edit.click()

            Cataloged_By_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='catalogedByDictionary']//span[@class='show-dictionary-popup dictionary-icon-enabled']"))
            )
            Cataloged_By_Edit.click()

            time.sleep(2)
            Staff_Select_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Connor Johnson']"))
            )
            Staff_Select_Edit.click()
            Collection_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[id='collectionDictionary'] span[class='show-dictionary-popup dictionary-icon-enabled']"))
            )
            Collection_Edit.click()

            time.sleep(2)

            Select_New_Record = Select(driver.find_element(By.CSS_SELECTOR, "div[class='dictionary-grid-container'] select[class='input-sm no-padding-left no-padding-right']"))
            Select_New_Record.select_by_visible_text('50')            #Selects the dropdown to make list bigger


            if excel_Collection.lower() == 'coffin family papers':
                Coffin_Family_Papers = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Coffin Family Papers']"))
                )
                Coffin_Family_Papers.click()

            if excel_Collection.lower() == 'college archives':
                College_Archives = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='College Archives']"))
                )
                College_Archives.click()

            if excel_Collection.lower() == 'college photo archives' or 'marist archive photograph collection':
                College_Photo_Archives = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='College Photo Archives']"))
                )
                College_Photo_Archives.click()

            if excel_Collection.lower() == 'honors thesis collection':
                Honors_Thesis_Collection = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Honors Thesis Collection']"))
                )
                Honors_Thesis_Collection.click()

            if excel_Collection.lower() == 'lowell thomas papers':
                Lowell_Thomas_Papers = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Lowell Thomas Papers']"))
                )
                Lowell_Thomas_Papers.click()

            if excel_Collection.lower() == 'lowell thomas papers - radio news show scripts':
                Lowell_Thomas_Papers_Radio = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Lowell Thomas Papers - Radio News Show Scripts']"))
                )
                Lowell_Thomas_Papers_Radio.click()

            if excel_Collection.lower() == 'oral histories':
                Oral_Histories = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Oral Histories']"))
                )
                Oral_Histories.click()

            if excel_Collection.lower() == 'poughkeepsie regatta collection':
                Poughkeepsie_Regatta_Collection = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Poughkeepsie Regatta Collection']"))
                )
                Poughkeepsie_Regatta_Collection.click()

            if excel_Collection.lower() == 'rare book collection':
                Rare_Book_Collection = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Rare Book Collection']"))
                )
                Rare_Book_Collection.click()

            if excel_Collection.lower() == 'robert hoe music collection':
                Robert_Hoe_Music_Collection = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Robert Hoe Music Collection']"))
                )
                Robert_Hoe_Music_Collection.click()

            if excel_Collection.lower() == 'student newspapers: the record and the circle':
                Student_Newspapers = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[normalize-space()='Student Newspapers: The Record and The Circle']"))
                )
                Student_Newspapers.click()

            if excel_Author.lower() != 'blank':
                time.sleep(2)
                dropdown_Click = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="photo"]/fieldset/div[1]/div[1]/div[1]/span/span[1]/span/span[2]'))
                )
                dropdown_Click.click()
                time.sleep(2)

                photographer_Input =WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[role='searchbox']"))
                )
                photographer_Input.send_keys(excel_Author)
                time.sleep(2)
                photographer_Input.send_keys(Keys.ENTER)



            time.sleep(4)
            Final_Save_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[data-bind='click: save, enable: isValid']"))
            )
            Final_Save_Edit.click()

            Back_Home = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='/']"))
            )
            Back_Home.click()

            Back_Catalogs = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#catalogs-home-button"))
            )
            Back_Catalogs.click()

        except:
            print("Could not finish collection Edit")

        row_num+=1

print("Task reached end")
time.sleep(4)
