#Connor Johnson
#Selenium bot to automate the photo portion inputs of data given by an excel sheet to the datebase: PastPerfect
#Good chrome extension to find paths in website: SelectorsHub

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from datetime import datetime

import time



# logging.basicConfig(filename='test.log', filemode='w', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


PATH = "/Users/Johnson_code/chromedriver"
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()
driver.get("https://mypastperfect.com/Account/Login?ReturnUrl=%2F")     #opens browser to website

username = driver.find_element(By.ID, 'Email')
username.send_keys("*********@marist.edu")                        #finds unsername

password = driver.find_element(By.ID, 'Password')
password.send_keys("********")
password.send_keys(Keys.RETURN)

             ###--------------Page Navigation--------------###

try:
    catalog_click = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "catalogs-home-button"))
    )
    catalog_click.click()#clicks the catalog button

    archives_click = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='#'][data-panel-class='archives']"))
    )
    archives_click.click()#clicks the photograph button

except:
    print("Could not find button")

            ###--------------Excel Import--------------###
book = load_workbook('PastPerfectOralHistoryUpload.xlsx')                               #excel file from personal computer
sheet = book.active                                                 #reads excel file inside folder
row_num = 2

while row_num <= sheet.max_row:
   for cell in sheet[row_num]:

        excel_Catalog = sheet['A' + str(row_num)].value                                #not using value yet, only tests
        excel_ObjectID = sheet['B' + str(row_num)].value
        excel_Object_Name = sheet['C' + str(row_num)].value
        excel_Title = sheet['D' + str(row_num)].value
        excel_Description = sheet['E' + str(row_num)].value.lstrip()
        excel_Collection = sheet['F' + str(row_num)].value
        excel_Date = sheet['G' + str(row_num)].value
        excel_Public_View = sheet['H' + str(row_num)].value
        excel_People = sheet['I' + str(row_num)].value
        excel_Subjects = sheet['J' + str(row_num)].value
        excel_Attachments = sheet['K' + str(row_num)].value
        excel_Image = sheet['L' + str(row_num)].value
        excel_Image_Title = sheet['M' + str(row_num)].value

                            ###--------------Insert New Record--------------###

        time.sleep(2)
        try:
            add_record_click = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="catalogs-grid"]/div/div[2]/div[1]/div/button[1]'))
            )
            add_record_click.click()  # clicks the 'add record' button

        except:
            print("Could not add new record")

        time.sleep(2)
        ObjectID_New_Record = driver.find_element(By.XPATH,'//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[4]/div/input')
        ObjectID_New_Record.send_keys(excel_ObjectID)  # identifier                    #finds Object ID Box and inserts from excel sheet

        Search_ObjectName_New_Record = driver.find_element(By.XPATH, '//*[@id="newRecordNameLex"]/div/div/span[2]')
        Search_ObjectName_New_Record.click()  # finds the search icon to go find Object Name

        time.sleep(2)
        ObjectName_New_Record = driver.find_element(By.XPATH, '/html/body/div[8]/div/div/div[2]/div/div[1]/div/div[1]/input')
        ObjectName_New_Record.send_keys(excel_Object_Name)  # Inserts object from excel into search

        time.sleep(2)
        Action = ActionChains(driver)
        Action_ObjectName_New_Record = driver.find_element(By.XPATH,"//td[normalize-space()='08: Communication Objects']")
        Action.double_click(Action_ObjectName_New_Record).perform()  # finds the 'Thesis' object name

        time.sleep(2)
        Title_New_Record = driver.find_element(By.XPATH,'//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[6]/div/input')
        Title_New_Record.send_keys(excel_Title)  # Inserts title from excel

        Description_New_Record = driver.find_element(By.XPATH,'//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[7]/div/textarea')
        Description_New_Record.send_keys(excel_Description)  # Inserts description from excel

        time.sleep(2)
        Add_New_Record = driver.find_element(By.XPATH,'//*[@id="newCatalogRecordDialog-modal"]/div/div/div[2]/div[3]/div[8]/button[1]')
        Add_New_Record.click()  # clicks the 'add new record' button
        time.sleep(2)




        ###--------------Going to second webpage after this point--------------###

        currentDate = time.strftime("%m/%d/%Y")
        Catalog_Date_Edit = driver.find_element(By.XPATH, '//*[@id="catalogDate"]/div/input')
        Catalog_Date_Edit.send_keys(currentDate)

        time.sleep(2)


        try:

            currentDate = time.strftime("%m/%d/%Y")
            Catalog_Date_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="catalogDate"]/div/input'))
            )
            Catalog_Date_Edit.send_keys(currentDate)

            date_formate = (excel_Date.strftime("%Y-%m-%d"))
            Date_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="Date"]'))
            )
            Date_Edit.send_keys(date_formate)

            if excel_Public_View.lower() == 'yes':
                Public_Access_Edit = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.NAME, 'public-access-checkbox'))
                )
                Public_Access_Edit.click()

            Cataloged_By_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[id='catalogedByDictionary'] span[class='show-dictionary-popup dictionary-icon-enabled']"))
            )
            Cataloged_By_Edit.click()

            Staff_Select_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "tbody tr:nth-child(6) td:nth-child(1)"))
            )
            Staff_Select_Edit.click()
            time.sleep(2)

            Collection_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[id='collectionDictionary'] span[class='show-dictionary-popup dictionary-icon-enabled']"))
            )
            Collection_Edit.click()

            time.sleep(2)

            ####oral history was not origionally on this so that is the next step you loser :)
            if excel_Collection.lower() == 'oral histories':
                Oral_History = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "tbody tr:nth-child(7) td:nth-child(1)"))
                )
                Oral_History.click()

        except:
            print("Could not finish Record Edit")


        if excel_Image.lower() != 'blank':
            try:
                Enter_Image_Managemnet = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.menu-btn-primary.col-md-8'))
                )
                Enter_Image_Managemnet.click()

                Add_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
                )
                Add_Image_Management.send_keys("/Users/connorjohnson/Desktop/For_PastPerfect/" + excel_Image)

                time.sleep(4)
                Checkbox_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "input[data-bind='checked: IsPublic, enable: $parent.editMode()']"))
                )
                Checkbox_Image_Management.click()

                Caption_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="image-management-page"]/div[5]/div[2]/div/input'))
                )
                Caption_Image_Management.send_keys(excel_Image_Title)

                Save_Image_Management = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "button[data-bind='click: save, enable: isSaveButtonEnabled()']"))
                )
                Save_Image_Management.click()
            except:
                print("Could not finish Image Management")

        # try:
        #
        #     driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        #
        #     Attach_URL_Dropdown = WebDriverWait(driver, 10).until(
        #         EC.presence_of_element_located((By.XPATH, "//body/div[@class='container']/div[@id='catalog-item']/div[@class='col-md-12 no-padding-left']/div[@id='ci-accordeon']/div[@id='attachments-tab']/div[@class='panel-heading']/h4[@class='panel-title']/span[1]"))
        #     )
        #     Attach_URL_Dropdown.click()
        #
        #     time.sleep(4)
        #
        #     Attachments_Split = excel_Attachments.split(';')
        #     for i in range(len(Attachments_Split)):
        #
        #         Add_Files = WebDriverWait(driver, 10).until(
        #             EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
        #         )
        #         #Add_Files.click()
        #         Add_Files.click
        #     # time.sleep(2)
        #
        #
        #     driver.execute_script("window.scrollTo(0,-350)")
        # except:
        #     print('Could not upload')

        try:
            time.sleep(4)

                        # -------------Works-----------#

            Final_Save_Edit = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[data-bind='click: save, enable: isValid']"))
            )
            Final_Save_Edit.click()

            Back_Home = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='/']"))
            )
            Back_Home.click()

            Back_Catalogs = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#catalogs-home-button"))
            )
            Back_Catalogs.click()

        except:
            print("Could not finish saving")



        row_num+=1



time.sleep(20)
# print("Task reached end")
